from openpyxl import Workbook
import pandas as pd

def export_CSV_Excel(masked_prompt_continuations, regard_metrics, toxicity, prompts, masks, regard_ratios):
    flat_list = []
    for row in regard_metrics:
         flat_list.extend(row)

    regard_flat = []
    for row in flat_list:
         regard_flat.extend(row)

    flat_cont = []
    for row in masked_prompt_continuations:
         flat_cont.extend(row)
    
    regard_positive = regard_flat[0::7]
    regard_negative = regard_flat[1::7]
    regard_neutral = regard_flat[2::7]
    regard_other = regard_flat[3::7]
    regard_difference = regard_flat[4::7]
    regard_difference1 = regard_flat[5::7]
    regard_difference05 = regard_flat[6::7]

    data = {'Completions':flat_cont, 'Regard Postive': regard_positive, 'Regard Negative':regard_negative, 'Regard Neutral': regard_neutral, 'Regard Other': regard_other, 'toxicity':toxicity, 'difference': regard_difference, 'diffrence-w-0.1-threshold': regard_difference1, 'diffrence-w-0.05-threshold': regard_difference05}

    df=pd.DataFrame(data)
    print (df)

    # Create combination of ratios dataframe
    combinations = [(prompt, mask) for prompt in prompts for mask in masks]

    dataRatios = []

    for i, (prompt, mask) in enumerate(combinations):
        row = [prompt, mask] + regard_ratios[i]
        dataRatios.append(row)


    columnsRatios = ['Prompt', 'Mask', 'Regard Positive Ratio', 'Regard Negative Ratio', 'Regard Neutral Ratio', 'Regard Other Ratio', 'Regard +/- Difference Ratio'] 
    dfRatios = pd.DataFrame(dataRatios, columns=columnsRatios)

    print(dfRatios)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Combined Data'


    # Print to csv
    for col_idx, col_name in enumerate(dfRatios.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)


    for r_idx, row in enumerate(dfRatios.iterrows(), start=2):
        for c_idx, value in enumerate(row[1], start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)


    start_row_df = len(dfRatios) + 3 


    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=start_row_df, column=col_idx, value=col_name)


    for r_idx, row in enumerate(df.iterrows(), start=start_row_df + 1):
        for c_idx, value in enumerate(row[1], start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)



    wb.save('BiasTool.xlsx')
