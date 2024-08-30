from openpyxl import Workbook
import pandas as pd

def export_CSV_Excel(masked_prompt_continuations, regard_metrics, llm_name):
    flat_list = []
    for row in regard_metrics:
         flat_list.extend(row)

    regard_flat = []
    for row in flat_list:
         regard_flat.extend(row)

    flat_cont = []
    for row in masked_prompt_continuations:
         flat_cont.extend(row)
    
    regard_positive = regard_flat[0::7]
    regard_negative = regard_flat[1::7]
    regard_neutral = regard_flat[2::7]
    regard_other = regard_flat[3::7]
    regard_difference = regard_flat[4::7]


    data = {'Completions':flat_cont, 'Regard Postive': regard_positive, 'Regard Negative':regard_negative, 'Regard Neutral': regard_neutral, 'Regard Other': regard_other, 'difference': regard_difference}

    df=pd.DataFrame(data)




    wb = Workbook()
    ws = wb.active
    ws.title = llm_name



    start_row_df = 1


    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=start_row_df, column=col_idx, value=col_name)


    for r_idx, row in enumerate(df.iterrows(), start=start_row_df + 1):
        for c_idx, value in enumerate(row[1], start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)



    wb.save('BiasTool.xlsx')
